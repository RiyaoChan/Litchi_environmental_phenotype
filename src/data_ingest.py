"""Read-only extraction; every value retains an Excel or Word locator."""
from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.utils.datetime import from_excel

from .io_utils import raw_paths, sha256, write_csv

EVENTS = {
    5: 'autumn_flush_mature', 6: 'white_tip', 7: 'inflorescence_emergence',
    8: 'initial_bloom', 9: 'full_bloom', 10: 'end_bloom',
    11: 'initial_fruit_set', 12: 'fruit_drop_1', 13: 'fruit_drop_2',
    14: 'preharvest_drop', 15: 'color_break', 16: 'maturity',
}
TRANSITIONS = {'P1': ('autumn_flush_mature', 'inflorescence_emergence'),
               'P2': ('inflorescence_emergence', 'full_bloom'),
               'P3': ('full_bloom', 'maturity')}
WEATHER_VARIABLES = ['tmean_c', 'tmin_c', 'tmax_c', 'precip_mm',
                     'relative_humidity_pct', 'sunshine_h', 'wind_speed']


def missing(value) -> bool:
    return value is None or (isinstance(value, str) and value.strip() in
                             {'', '-', '—', 'NA', 'N/A', 'na'})


def number(value) -> float:
    if missing(value) or isinstance(value, bool):
        return np.nan
    try:
        result = float(value)
        return result if np.isfinite(result) else np.nan
    except (TypeError, ValueError):
        return np.nan


def raw_text(value) -> str:
    if value is None:
        return ''
    return value.isoformat() if isinstance(value, (date, datetime)) else str(value)


def merged_value(sheet, row, column):
    cell = sheet.cell(row, column)
    if cell.value is not None:
        return cell.value, cell.coordinate
    for rng in sheet.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= column <= rng.max_col:
            anchor = sheet.cell(rng.min_row, rng.min_col)
            return anchor.value, anchor.coordinate
    return None, cell.coordinate


def identify_orchard(value, cfg):
    matches = [k for k, v in cfg['orchards'].items() if v['name'] in str(value)]
    if len(matches) != 1:
        raise ValueError(f'Unknown/ambiguous orchard label: {value!r}')
    return matches[0]


def parse_full_date(value, epoch, accept_serials=False):
    """Never attach a harvest year to month/day. Untyped serials are audit-only."""
    if missing(value):
        return pd.NaT, pd.NaT, 'missing'
    if isinstance(value, (date, datetime)):
        parsed = pd.Timestamp(value).normalize()
        return parsed, parsed, 'ok'
    if isinstance(value, (int, float)):
        if not np.isfinite(value) or value != int(value):
            return pd.NaT, pd.NaT, 'invalid_serial'
        try:
            candidate = pd.Timestamp(from_excel(value, epoch)).normalize()
        except (ValueError, TypeError, OverflowError):
            return pd.NaT, pd.NaT, 'invalid_serial'
        if accept_serials:
            return candidate, candidate, 'ok'
        return pd.NaT, candidate, 'unformatted_excel_serial_requires_confirmation'
    text = str(value).strip()
    match = re.fullmatch(r'(\d{4})[-/年.](\d{1,2})[-/月.](\d{1,2})日?(?:[ T]00:00:00)?', text)
    if match:
        try:
            parsed = pd.Timestamp(date(*map(int, match.groups())))
            return parsed, parsed, 'ok'
        except ValueError:
            return pd.NaT, pd.NaT, 'invalid_complete_date'
    if re.fullmatch(r'\d{1,2}[-/月.]\d{1,2}日?', text):
        return pd.NaT, pd.NaT, 'year_missing'
    return pd.NaT, pd.NaT, 'qualitative_not_date'


def normalize_season_date(value, epoch, harvest_year, event_name, constraints):
    """User-authorized month/day extraction with A-column harvest-year authority.

    Original decoded years are retained for provenance, not treated as observed
    phenology years. No month/day is shifted to enforce sequence.
    """
    parsed, candidate, flag = parse_full_date(value, epoch, accept_serials=True)
    if not constraints.get('seasonal_year_normalization'):
        return (*parse_full_date(value, epoch, constraints.get('accept_unformatted_excel_serials', False)), 'original_complete_date')
    if constraints.get('date_year_authority') != 'harvest_year_column_A':
        raise ValueError('Seasonal normalization requires explicit A-column authority')
    if pd.isna(candidate):
        match = re.fullmatch(r'(\d{1,2})[-/月.](\d{1,2})日?', str(value).strip())
        if not match:
            return pd.NaT, candidate, flag, 'not_a_date'
        month, day = map(int, match.groups())
    else:
        month, day = candidate.month, candidate.day
    target_year = int(harvest_year) - int(month >= constraints['season_start_month'])
    if event_name == 'autumn_flush_mature' and target_year != int(harvest_year) - 1:
        return pd.NaT, candidate, 'autumn_month_inconsistent_with_previous_year', 'user_authorized_season_rule'
    try:
        normalized = pd.Timestamp(date(target_year, month, day))
    except ValueError:
        return pd.NaT, candidate, 'invalid_rebased_calendar_date', 'user_authorized_season_rule'
    return normalized, candidate, 'ok', 'A_column_harvest_year_plus_original_month_day'


def read_phenology(root, cfg):
    path = root / cfg['inputs']['phenology']
    wb = openpyxl.load_workbook(path, data_only=True)
    records = []
    try:
        for sheet in wb:
            for row in range(2, sheet.max_row + 1):
                if missing(sheet.cell(row, 2).value):
                    continue
                orchard = identify_orchard(sheet.cell(row, 2).value, cfg)
                meta = cfg['orchards'][orchard]
                year_raw, year_cell = merged_value(sheet, row, 1)
                year = int(year_raw)
                for column, event in EVENTS.items():
                    cell = sheet.cell(row, column)
                    parsed, candidate, flag, derivation = normalize_season_date(
                        cell.value, wb.epoch, year, event, cfg['constraints'])
                    flags = [] if flag == 'ok' else [flag]
                    if pd.notna(parsed) and parsed.year not in (year - 1, year):
                        flags.append('normalized_year_outside_harvest_season')
                        parsed = pd.NaT
                    if event == 'maturity' and pd.notna(parsed) and parsed.year != year:
                        flags.append('maturity_year_disagrees_with_harvest_year')
                        parsed = pd.NaT
                    records.append({
                        'orchard_id': orchard, 'region_id': meta['region_id'],
                        'station_id': meta['station_id'], 'cultivar': sheet.cell(row, 3).value,
                        'harvest_year': year, 'source_block_id': f'pheno_{sheet.title}_r{row:02}',
                        'event_name': event, 'event_date': parsed,
                        'calendar_year': parsed.year if pd.notna(parsed) else np.nan,
                        'day_of_year': parsed.dayofyear if pd.notna(parsed) else np.nan,
                        'decoded_candidate_date': candidate,
                        'date_derivation': derivation,
                        'date_year_rebased': int(pd.notna(parsed) and pd.notna(candidate) and parsed.year != candidate.year),
                        'date_authority': cfg['constraints'].get('date_year_authority','original_full_date'),
                        'date_authority_source': cfg['constraints'].get('user_amendment',''),
                        'source_file': path.name, 'source_sheet': sheet.title,
                        'source_row': row, 'source_column': cell.column_letter,
                        'source_cell': cell.coordinate, 'harvest_year_source_cell': year_cell,
                        'raw_value': raw_text(cell.value), 'cell_data_type': cell.data_type,
                        'number_format': cell.number_format,
                        'qc_flag': ';'.join(flags) or 'ok',
                    })
    finally:
        wb.close()
    return pd.DataFrame(records)


def read_yield(root, cfg):
    path = root / cfg['inputs']['yield']
    wb = openpyxl.load_workbook(path, data_only=True)
    records = []
    a39 = wb['Sheet1']['A39'].value
    conflict = a39 != cfg['constraints']['a39_expected']
    try:
        for sheet in wb:
            for row in range(3, sheet.max_row + 1):
                if missing(sheet.cell(row, 3).value):
                    continue
                orchard_label, orchard_cell = merged_value(sheet, row, 2)
                orchard = identify_orchard(orchard_label, cfg)
                year_raw, year_cell = merged_value(sheet, row, 1)
                meta = cfg['orchards'][orchard]
                block_start = int(re.search(r'\d+', orchard_cell).group())
                record = {
                    'orchard_id': orchard, 'region_id': meta['region_id'],
                    'station_id': meta['station_id'], 'cultivar': meta['cultivar'],
                    'harvest_year': int(year_raw),
                    'source_block_id': f'yield_{sheet.title}_r{block_start:02}',
                    'plot_id': None, 'tree_id': None,
                    'tree_class': str(sheet.cell(row, 3).value).strip(),
                    'source_file': path.name, 'source_sheet': sheet.title,
                    'source_row': row, 'year_source_cell': year_cell,
                    'year_protocol_conflict': int(conflict and year_cell == 'A39'),
                }
                fields = {4: 'tree_count', 5: 'sample_tree_1_yield_kg',
                          6: 'sample_tree_2_yield_kg', 7: 'sample_tree_3_yield_kg',
                          8: 'single_fruit_weight_g', 9: 'diseased_pest_fruit_pct',
                          10: 'reported_class_mean_yield_kg', 11: 'reported_yield_kg_per_mu'}
                for col, name in fields.items():
                    value, cell = merged_value(sheet, row, col) if col == 11 else (sheet.cell(row, col).value, sheet.cell(row, col).coordinate)
                    record[name] = number(value)
                    record[name + '_source_cell'] = cell
                    record[name + '_raw'] = raw_text(value)
                sample = [record[f'sample_tree_{i}_yield_kg'] for i in (1, 2, 3)]
                record['sample_tree_mean_yield_kg'] = float(np.mean(sample)) if all(np.isfinite(sample)) else np.nan
                record['class_mean_difference_kg'] = record['reported_class_mean_yield_kg'] - record['sample_tree_mean_yield_kg']
                record['class_mean_mismatch'] = int(abs(record['class_mean_difference_kg']) > cfg['qc']['class_mean_tolerance_kg'])
                records.append(record)
    finally:
        wb.close()
    return pd.DataFrame(records), {'actual_a39': a39, 'expected_a39': cfg['constraints']['a39_expected'], 'conflict': conflict}


def weather_header(value):
    text = str(value or '')
    mappings = [('最高气温', 'tmax_c'), ('最低气温', 'tmin_c'),
                ('平均气温', 'tmean_c'), ('降水', 'precip_mm'),
                ('湿度', 'relative_humidity_pct'), ('日照', 'sunshine_h'),
                ('风速', 'wind_speed')]
    return next((target for key, target in mappings if key in text), None)


def read_weather(root, cfg):
    records, sheets, rejected = [], [], []
    for path in sorted((root / cfg['inputs']['weather_dir']).glob('*.xlsx')):
        if path.name.startswith('~$'):
            continue
        region = 'haikou' if '海口' in path.name else 'lingshui' if '陵水' in path.name else None
        if region is None:
            raise ValueError(f'Unmapped weather source: {path.name}')
        wb = openpyxl.load_workbook(path, data_only=True)
        try:
            for sheet in wb:
                columns = {i: weather_header(c.value) for i, c in enumerate(sheet[1], 1) if weather_header(c.value)}
                sheets.append({'source_file': path.relative_to(root).as_posix(),
                               'source_sheet': sheet.title,
                               'column_mapping': json.dumps({sheet.cell(1,i).coordinate: {'header': sheet.cell(1,i).value,'variable':v} for i,v in columns.items()}, ensure_ascii=False)})
                for row in range(2, sheet.max_row + 1):
                    if all(missing(c.value) for c in sheet[row]):
                        continue
                    parts = [sheet.cell(row, c).value for c in (2, 3, 4)]
                    try:
                        if not all(isinstance(x, (int, float)) and x == int(x) for x in parts):
                            raise ValueError('missing/noninteger date component')
                        day = pd.Timestamp(date(*map(int, parts)))
                    except (ValueError, TypeError, OverflowError) as exc:
                        rejected.append({'source_file': path.relative_to(root).as_posix(), 'source_sheet': sheet.title, 'source_row': row, 'reason': str(exc), 'raw_row': json.dumps([raw_text(c.value) for c in sheet[row]], ensure_ascii=False)})
                        continue
                    rec = {'station_id': region + '_region_proxy', 'region_id': region,
                           'source_station_code': None, 'date': day,
                           'source_file': path.relative_to(root).as_posix(),
                           'source_sheet': sheet.title, 'source_row': row}
                    for var in WEATHER_VARIABLES:
                        rec[var] = np.nan
                        rec[var + '_raw'] = ''
                        rec[var + '_source_cell'] = ''
                    for col, var in columns.items():
                        rec[var] = number(sheet.cell(row, col).value)
                        rec[var + '_raw'] = raw_text(sheet.cell(row, col).value)
                        rec[var + '_source_cell'] = sheet.cell(row, col).coordinate
                    records.append(rec)
        finally:
            wb.close()
    return pd.DataFrame(records), pd.DataFrame(sheets), pd.DataFrame(rejected, columns=['source_file','source_sheet','source_row','reason','raw_row'])


def read_document_evidence(root, cfg):
    documents, evidence, cells = [], [], []
    for path in [p for p in raw_paths(root, cfg) if p.suffix == '.doc']:
        digest = sha256(path)
        cache = root / 'data/interim/documents' / (digest + '.json')
        if not cache.exists():
            raise RuntimeError('Word extraction cache missing. Run python scripts/extract_word_sources.py before audit.')
        data = json.loads(cache.read_text(encoding='utf-8'))
        if data['source_sha256'] != digest:
            raise RuntimeError('Word cache/source hash mismatch')
        source = path.relative_to(root).as_posix()
        documents.append({'source_file': source, 'sha256': digest, 'characters':len(data['text']), 'table_count':len(data['tables'])})
        for line_num, line in enumerate(data['text'].splitlines(), 1):
            if any(k in line for k in ['台风', '放丢投产', '没有投产', '树体恢复', '树冠还在恢复']):
                evidence.append({'source_file': source, 'text_line':line_num, 'text':line})
        for table in data['tables']:
            for cell in table['cells']:
                cells.append({'source_file': source, 'table_index':table['table_index'], **cell})
    write_csv(root / 'results/inventory/document_inventory.csv', documents)
    write_csv(root / 'results/qc/document_damage_evidence.csv', evidence)
    write_csv(root / 'data/interim/annual_document_table_cells.csv', cells)
    return pd.DataFrame(documents), pd.DataFrame(evidence), pd.DataFrame(cells)
